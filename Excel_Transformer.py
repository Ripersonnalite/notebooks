import os
import csv
import time
import schedule
import datetime
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

def processa():
    try:
        # Adicionar entrada ao log indicando o início da execução
        # Adicionar entrada ao log
        #with open('runner_log.txt', 'a') as log_file:
        #    log_file.write(f"{datetime.now()} - Script '{script_name}' executando.\n")
        # Caminho para o arquivo Excel original
        excel_path = r'C:\Users\FLSV\Origin.xlsx'
        # Caminho para o arquivo Excel de destino
        destination_excel_path = r'C:\Users\FLSV\Destination.xlsx'

        # Carregar o workbook e a planilha específica
        workbook = load_workbook(excel_path, data_only=True)
        worksheet = workbook['RTI']

        # Acessar a tabela nomeada
        table = worksheet.tables['con_67_ys']

        # Obter o intervalo da tabela
        table_range = table.ref

        # Extrair os dados da tabela
        data = worksheet[table_range]

        # Converter os dados em uma lista de dicionários
        data_list = []
        columns = [cell.value for cell in data[0]]  # Primeira linha para nomes das colunas

        for row in data[1:]:
            row_data = {columns[i]: row[i].value for i in range(len(columns))}
            data_list.append(row_data)

        # Converter a lista de dicionários em um DataFrame
        df = pd.DataFrame(data_list)

        # Remover espaços em branco no início/fim dos nomes das colunas
        df.columns = df.columns.str.strip()

        # Filtrar os dados onde 'Status Prazo' == 'Vencida'
        filtered_df = df[df['Status Prazo'] == 'Vencida']

        # Contar as ocorrências de cada tipo de 'Plataforma' quando 'Status Prazo' é 'Vencida'
        summary = filtered_df['Plataforma'].value_counts().reset_index()
        summary.columns = ['Plataforma', 'Count']

        # Adicionar uma coluna com a data atual
        current_date = datetime.now().strftime('%#m/%#d/%Y')
        summary['Data de Processamento'] = current_date
        print(summary)

        # Carregar o workbook de destino e a planilha específica
        destination_workbook = load_workbook(destination_excel_path)
        destination_worksheet = destination_workbook['Sheet1']

        # Acessar a tabela nomeada
        destination_table = destination_worksheet.tables['Table1']

        # Obter o intervalo da tabela
        destination_table_range = destination_table.ref

        # Extrair os dados da tabela de destino
        destination_data = destination_worksheet[destination_table_range]

        # Converter os dados da tabela de destino em um DataFrame
        destination_data_list = []
        destination_columns = [cell.value for cell in destination_data[0]]  # Primeira linha para nomes das colunas

        for row in destination_data[1:]:
            row_data = {destination_columns[i]: row[i].value for i in range(len(destination_columns))}
            destination_data_list.append(row_data)

        # Converter a lista de dicionários em um DataFrame
        destination_df = pd.DataFrame(destination_data_list)

        # Remover espaços em branco no início/fim dos nomes das colunas
        destination_df.columns = destination_df.columns.str.strip()

        # Verificar se já existem dados com a mesma 'Data de Processamento'
        existing_dates = destination_df['Data de Processamento'] == current_date

        # Se existirem, remover essas linhas
        if existing_dates.any():
            destination_df = destination_df[~existing_dates]

        # Concatenar o novo resumo com os dados existentes
        updated_df = pd.concat([destination_df, summary], ignore_index=True)

        # Limpar os dados antigos da tabela no Excel
        for row in destination_worksheet[destination_table_range]:
            for cell in row:
                cell.value = None

        # Escrever os novos dados na planilha
        for r_idx, row in enumerate(dataframe_to_rows(updated_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                destination_worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Atualizar o intervalo da tabela para incluir os novos dados
        new_table_range = f"A1:{chr(64 + len(updated_df.columns))}{len(updated_df) + 1}"
        destination_table.ref = new_table_range

        # Salvar o workbook
        destination_workbook.save(destination_excel_path)

        # Nome do script para o log
        script_name = os.path.basename(__file__)

        # Adicionar entrada ao log
        with open('runner_log.txt', 'a') as log_file:
            log_file.write(f"{datetime.now()} - Script '{script_name}' executado.\n")
    except Exception as e:
        # Capturar qualquer exceção que ocorra durante a execução
        with open('runner_log.txt', 'a') as log_file:
            log_file.write(f"{datetime.now()} - Erro durante a execução do script '{script_name}': {e}\n")
        


# Schedule the script to run at the specified times on the specified days
try:
    schedule.every().monday.at("09:00").do(processa)
    schedule.every().monday.at("09:30").do(processa)
    schedule.every().monday.at("10:00").do(processa)

    schedule.every().wednesday.at("09:00").do(processa)
    schedule.every().wednesday.at("09:30").do(processa)
    schedule.every().wednesday.at("10:00").do(processa)

    schedule.every().thursday.at("09:00").do(processa)
    schedule.every().thursday.at("09:30").do(processa)
    schedule.every().thursday.at("13:45").do(processa)

    schedule.every().friday.at("09:00").do(processa)
    schedule.every().friday.at("09:30").do(processa)
    schedule.every().friday.at("10:00").do(processa)
except Exception as e:
        # Capturar qualquer exceção que ocorra durante a execução
        with open('runner_log.txt', 'a') as log_file:
            log_file.write(f"{datetime.now()} - Erro durante a execução do script '{script_name}': {e}\n")
 
   

if __name__ == "__main__":
    processa()
    #while True:
    #    schedule.run_pending()
    #    time.sleep(1) 